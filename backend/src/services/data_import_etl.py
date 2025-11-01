"""
数据导入ETL逻辑
处理复杂文档格式，支持多种数据源导入
"""

import asyncio
import pandas as pd
import numpy as np
from typing import Dict, List, Any, Optional, Union, Tuple
from datetime import datetime, timedelta
import logging
import json
import re
from pathlib import Path
import openpyxl
import csv
import sqlite3
from dataclasses import dataclass, asdict
from enum import Enum

logger = logging.getLogger(__name__)

class DataSourceType(Enum):
    """数据源类型"""
    EXCEL = "excel"
    CSV = "csv"
    JSON = "json"
    XML = "xml"
    DATABASE = "database"
    API = "api"

class DocumentFormat(Enum):
    """文档格式类型"""
    SINGLE_HEADER_MULTI_ROWS = "single_header_multi_rows"  # 标准格式
    MULTI_HEADER_MULTI_ROWS = "multi_header_multi_rows"    # 多行表头
    HEADER_FIRST_ROW_ONLY = "header_first_row_only"        # 只有第一行有表头
    SEPARATE_HEADER_BODY = "separate_header_body"          # 表头和表体分离
    HEADER_ONLY = "header_only"                            # 只有表头
    BODY_ONLY = "body_only"                                # 只有表体

class DataQualityLevel(Enum):
    """数据质量等级"""
    EXCELLENT = "excellent"  # 95%+
    GOOD = "good"           # 85-95%
    FAIR = "fair"           # 70-85%
    POOR = "poor"           # <70%

@dataclass
class ImportResult:
    """导入结果"""
    success: bool
    total_records: int
    successful_records: int
    failed_records: int
    quality_score: float
    quality_level: DataQualityLevel
    errors: List[str]
    warnings: List[str]
    processing_time: float
    import_id: str
    timestamp: datetime

@dataclass
class FieldMapping:
    """字段映射"""
    source_field: str
    target_field: str
    data_type: str
    transformation_rule: Optional[str] = None
    validation_rule: Optional[str] = None
    is_required: bool = False

@dataclass
class DataValidationRule:
    """数据验证规则"""
    field_name: str
    rule_type: str  # 'range', 'format', 'enum', 'custom'
    rule_value: Any
    error_message: str

class DataImportETL:
    """数据导入ETL处理器"""
    
    def __init__(self, db_service, cache_service):
        self.db_service = db_service
        self.cache_service = cache_service
        
        # 数据质量检查器
        self.quality_checker = DataQualityChecker()
        
        # 字段映射器
        self.field_mapper = FieldMapper()
        
        # 数据转换器
        self.data_transformer = DataTransformer()
        
        # 验证规则
        self.validation_rules = self._load_validation_rules()
    
    async def process_data_import(
        self,
        file_path: str,
        source_type: DataSourceType,
        document_format: DocumentFormat,
        field_mappings: List[FieldMapping],
        target_table: str,
        import_config: Dict[str, Any]
    ) -> ImportResult:
        """处理数据导入"""
        start_time = datetime.now()
        import_id = f"import_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        
        try:
            # 1. 读取原始数据
            raw_data = await self._read_source_data(file_path, source_type)
            
            # 2. 检测文档格式
            detected_format = await self._detect_document_format(raw_data)
            if detected_format != document_format:
                logger.warning(f"检测到的格式 {detected_format} 与指定格式 {document_format} 不匹配")
            
            # 3. 解析数据结构
            parsed_data = await self._parse_document_structure(raw_data, document_format)
            
            # 4. 数据质量检查
            quality_report = await self._check_data_quality(parsed_data)
            
            # 5. 字段映射和转换
            mapped_data = await self._apply_field_mappings(parsed_data, field_mappings)
            
            # 6. 数据验证
            validation_result = await self._validate_data(mapped_data, field_mappings)
            
            # 7. 数据清洗
            cleaned_data = await self._clean_data(mapped_data, validation_result)
            
            # 8. 数据加载
            load_result = await self._load_data_to_target(cleaned_data, target_table, import_config)
            
            # 9. 生成导入结果
            processing_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=load_result["success"],
                total_records=len(parsed_data),
                successful_records=load_result["successful_count"],
                failed_records=load_result["failed_count"],
                quality_score=quality_report["overall_score"],
                quality_level=quality_report["quality_level"],
                errors=validation_result["errors"],
                warnings=quality_report["warnings"],
                processing_time=processing_time,
                import_id=import_id,
                timestamp=datetime.now()
            )
            
        except Exception as e:
            logger.error(f"Data import failed: {str(e)}")
            processing_time = (datetime.now() - start_time).total_seconds()
            
            return ImportResult(
                success=False,
                total_records=0,
                successful_records=0,
                failed_records=0,
                quality_score=0.0,
                quality_level=DataQualityLevel.POOR,
                errors=[str(e)],
                warnings=[],
                processing_time=processing_time,
                import_id=import_id,
                timestamp=datetime.now()
            )
    
    async def _read_source_data(self, file_path: str, source_type: DataSourceType) -> Any:
        """读取源数据"""
        try:
            if source_type == DataSourceType.EXCEL:
                return await self._read_excel_file(file_path)
            elif source_type == DataSourceType.CSV:
                return await self._read_csv_file(file_path)
            elif source_type == DataSourceType.JSON:
                return await self._read_json_file(file_path)
            elif source_type == DataSourceType.XML:
                return await self._read_xml_file(file_path)
            elif source_type == DataSourceType.DATABASE:
                return await self._read_database_data(file_path)
            elif source_type == DataSourceType.API:
                return await self._read_api_data(file_path)
            else:
                raise ValueError(f"不支持的数据源类型: {source_type}")
                
        except Exception as e:
            logger.error(f"Failed to read source data: {str(e)}")
            raise
    
    async def _read_excel_file(self, file_path: str) -> Dict[str, Any]:
        """读取Excel文件"""
        try:
            workbook = openpyxl.load_workbook(file_path, data_only=True)
            sheets_data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                
                # 读取所有数据
                data = []
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):  # 跳过空行
                        data.append(list(row))
                
                sheets_data[sheet_name] = {
                    "data": data,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column
                }
            
            return {
                "type": "excel",
                "file_path": file_path,
                "sheets": sheets_data
            }
            
        except Exception as e:
            logger.error(f"Failed to read Excel file: {str(e)}")
            raise
    
    async def _read_csv_file(self, file_path: str) -> Dict[str, Any]:
        """读取CSV文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                # 检测编码
                content = file.read()
                
                # 尝试不同的分隔符
                separators = [',', ';', '\t', '|']
                detected_separator = ','
                
                for sep in separators:
                    if content.count(sep) > content.count(detected_separator):
                        detected_separator = sep
                
                # 重新读取文件
                file.seek(0)
                reader = csv.reader(file, delimiter=detected_separator)
                data = list(reader)
            
            return {
                "type": "csv",
                "file_path": file_path,
                "data": data,
                "separator": detected_separator
            }
            
        except Exception as e:
            logger.error(f"Failed to read CSV file: {str(e)}")
            raise
    
    async def _read_json_file(self, file_path: str) -> Dict[str, Any]:
        """读取JSON文件"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                data = json.load(file)
            
            return {
                "type": "json",
                "file_path": file_path,
                "data": data
            }
            
        except Exception as e:
            logger.error(f"Failed to read JSON file: {str(e)}")
            raise
    
    async def _read_xml_file(self, file_path: str) -> Dict[str, Any]:
        """读取XML文件"""
        try:
            import xml.etree.ElementTree as ET
            
            tree = ET.parse(file_path)
            root = tree.getroot()
            
            # 将XML转换为字典
            def xml_to_dict(element):
                result = {}
                
                # 添加属性
                if element.attrib:
                    result['@attributes'] = element.attrib
                
                # 添加文本内容
                if element.text and element.text.strip():
                    result['text'] = element.text.strip()
                
                # 添加子元素
                for child in element:
                    child_dict = xml_to_dict(child)
                    if child.tag in result:
                        if not isinstance(result[child.tag], list):
                            result[child.tag] = [result[child.tag]]
                        result[child.tag].append(child_dict)
                    else:
                        result[child.tag] = child_dict
                
                return result
            
            return {
                "type": "xml",
                "file_path": file_path,
                "data": xml_to_dict(root)
            }
            
        except Exception as e:
            logger.error(f"Failed to read XML file: {str(e)}")
            raise
    
    async def _read_database_data(self, connection_string: str) -> Dict[str, Any]:
        """读取数据库数据"""
        try:
            # 这里可以实现从各种数据库读取数据
            # 支持PostgreSQL, MySQL, SQLite等
            
            if connection_string.startswith('sqlite'):
                conn = sqlite3.connect(connection_string.replace('sqlite://', ''))
                cursor = conn.cursor()
                
                # 获取表列表
                cursor.execute("SELECT name FROM sqlite_master WHERE type='table';")
                tables = [row[0] for row in cursor.fetchall()]
                
                data = {}
                for table in tables:
                    cursor.execute(f"SELECT * FROM {table}")
                    data[table] = cursor.fetchall()
                
                conn.close()
                
                return {
                    "type": "database",
                    "connection_string": connection_string,
                    "tables": data
                }
            
            else:
                raise ValueError(f"不支持的数据库类型: {connection_string}")
                
        except Exception as e:
            logger.error(f"Failed to read database data: {str(e)}")
            raise
    
    async def _read_api_data(self, api_endpoint: str) -> Dict[str, Any]:
        """读取API数据"""
        try:
            import aiohttp
            
            async with aiohttp.ClientSession() as session:
                async with session.get(api_endpoint) as response:
                    if response.status == 200:
                        data = await response.json()
                        return {
                            "type": "api",
                            "endpoint": api_endpoint,
                            "data": data
                        }
                    else:
                        raise Exception(f"API请求失败: {response.status}")
                        
        except Exception as e:
            logger.error(f"Failed to read API data: {str(e)}")
            raise
    
    async def _detect_document_format(self, raw_data: Dict[str, Any]) -> DocumentFormat:
        """检测文档格式"""
        try:
            if raw_data["type"] == "excel":
                return await self._detect_excel_format(raw_data)
            elif raw_data["type"] == "csv":
                return await self._detect_csv_format(raw_data)
            else:
                return DocumentFormat.SINGLE_HEADER_MULTI_ROWS
                
        except Exception as e:
            logger.error(f"Failed to detect document format: {str(e)}")
            return DocumentFormat.SINGLE_HEADER_MULTI_ROWS
    
    async def _detect_excel_format(self, excel_data: Dict[str, Any]) -> DocumentFormat:
        """检测Excel格式"""
        try:
            # 获取第一个工作表
            first_sheet = list(excel_data["sheets"].values())[0]
            data = first_sheet["data"]
            
            if not data:
                return DocumentFormat.HEADER_ONLY
            
            # 分析前几行数据
            header_row = data[0]
            has_header = any(isinstance(cell, str) and cell.strip() for cell in header_row)
            
            if not has_header:
                return DocumentFormat.BODY_ONLY
            
            # 检查是否有多个表头行
            header_rows = 1
            for i in range(1, min(5, len(data))):
                row = data[i]
                if any(isinstance(cell, str) and cell.strip() and not str(cell).replace('.', '').replace('-', '').isdigit() for cell in row):
                    header_rows += 1
                else:
                    break
            
            if header_rows > 1:
                return DocumentFormat.MULTI_HEADER_MULTI_ROWS
            else:
                return DocumentFormat.SINGLE_HEADER_MULTI_ROWS
                
        except Exception as e:
            logger.error(f"Failed to detect Excel format: {str(e)}")
            return DocumentFormat.SINGLE_HEADER_MULTI_ROWS
    
    async def _detect_csv_format(self, csv_data: Dict[str, Any]) -> DocumentFormat:
        """检测CSV格式"""
        try:
            data = csv_data["data"]
            
            if not data:
                return DocumentFormat.HEADER_ONLY
            
            # 分析第一行
            first_row = data[0]
            has_header = any(isinstance(cell, str) and cell.strip() for cell in first_row)
            
            if not has_header:
                return DocumentFormat.BODY_ONLY
            
            return DocumentFormat.SINGLE_HEADER_MULTI_ROWS
            
        except Exception as e:
            logger.error(f"Failed to detect CSV format: {str(e)}")
            return DocumentFormat.SINGLE_HEADER_MULTI_ROWS
    
    async def _parse_document_structure(
        self, 
        raw_data: Dict[str, Any], 
        document_format: DocumentFormat
    ) -> Dict[str, Any]:
        """解析文档结构"""
        try:
            if raw_data["type"] == "excel":
                return await self._parse_excel_structure(raw_data, document_format)
            elif raw_data["type"] == "csv":
                return await self._parse_csv_structure(raw_data, document_format)
            elif raw_data["type"] == "json":
                return await self._parse_json_structure(raw_data, document_format)
            else:
                raise ValueError(f"不支持的数据类型: {raw_data['type']}")
                
        except Exception as e:
            logger.error(f"Failed to parse document structure: {str(e)}")
            raise
    
    async def _parse_excel_structure(
        self, 
        excel_data: Dict[str, Any], 
        document_format: DocumentFormat
    ) -> Dict[str, Any]:
        """解析Excel结构"""
        try:
            parsed_sheets = {}
            
            for sheet_name, sheet_data in excel_data["sheets"].items():
                data = sheet_data["data"]
                
                if document_format == DocumentFormat.SINGLE_HEADER_MULTI_ROWS:
                    if len(data) > 0:
                        headers = data[0]
                        rows = data[1:] if len(data) > 1 else []
                        
                        parsed_sheets[sheet_name] = {
                            "headers": headers,
                            "rows": rows,
                            "format": "single_header_multi_rows"
                        }
                
                elif document_format == DocumentFormat.MULTI_HEADER_MULTI_ROWS:
                    # 检测表头行数
                    header_rows = self._detect_header_rows(data)
                    headers = self._merge_header_rows(data[:header_rows])
                    rows = data[header_rows:]
                    
                    parsed_sheets[sheet_name] = {
                        "headers": headers,
                        "rows": rows,
                        "format": "multi_header_multi_rows",
                        "header_rows": header_rows
                    }
                
                elif document_format == DocumentFormat.HEADER_ONLY:
                    parsed_sheets[sheet_name] = {
                        "headers": data[0] if data else [],
                        "rows": [],
                        "format": "header_only"
                    }
                
                elif document_format == DocumentFormat.BODY_ONLY:
                    parsed_sheets[sheet_name] = {
                        "headers": [],
                        "rows": data,
                        "format": "body_only"
                    }
            
            return {
                "type": "parsed_excel",
                "sheets": parsed_sheets,
                "original_format": document_format
            }
            
        except Exception as e:
            logger.error(f"Failed to parse Excel structure: {str(e)}")
            raise
    
    async def _parse_csv_structure(
        self, 
        csv_data: Dict[str, Any], 
        document_format: DocumentFormat
    ) -> Dict[str, Any]:
        """解析CSV结构"""
        try:
            data = csv_data["data"]
            
            if document_format == DocumentFormat.SINGLE_HEADER_MULTI_ROWS:
                headers = data[0] if data else []
                rows = data[1:] if len(data) > 1 else []
                
                return {
                    "type": "parsed_csv",
                    "headers": headers,
                    "rows": rows,
                    "format": "single_header_multi_rows"
                }
            
            elif document_format == DocumentFormat.HEADER_ONLY:
                return {
                    "type": "parsed_csv",
                    "headers": data[0] if data else [],
                    "rows": [],
                    "format": "header_only"
                }
            
            elif document_format == DocumentFormat.BODY_ONLY:
                return {
                    "type": "parsed_csv",
                    "headers": [],
                    "rows": data,
                    "format": "body_only"
                }
            
            else:
                raise ValueError(f"不支持的CSV格式: {document_format}")
                
        except Exception as e:
            logger.error(f"Failed to parse CSV structure: {str(e)}")
            raise
    
    async def _parse_json_structure(
        self, 
        json_data: Dict[str, Any], 
        document_format: DocumentFormat
    ) -> Dict[str, Any]:
        """解析JSON结构"""
        try:
            data = json_data["data"]
            
            # 检测JSON结构类型
            if isinstance(data, list):
                if data and isinstance(data[0], dict):
                    # 数组对象格式
                    headers = list(data[0].keys()) if data else []
                    rows = [list(row.values()) for row in data]
                    
                    return {
                        "type": "parsed_json",
                        "headers": headers,
                        "rows": rows,
                        "format": "array_objects"
                    }
                else:
                    # 简单数组格式
                    return {
                        "type": "parsed_json",
                        "headers": [],
                        "rows": data,
                        "format": "simple_array"
                    }
            
            elif isinstance(data, dict):
                # 对象格式，尝试转换为表格
                return await self._convert_dict_to_table(data)
            
            else:
                raise ValueError(f"不支持的JSON格式: {type(data)}")
                
        except Exception as e:
            logger.error(f"Failed to parse JSON structure: {str(e)}")
            raise
    
    async def _convert_dict_to_table(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """将字典转换为表格格式"""
        try:
            # 扁平化嵌套字典
            flattened = self._flatten_dict(data)
            
            headers = list(flattened.keys())
            rows = [list(flattened.values())]
            
            return {
                "type": "parsed_json",
                "headers": headers,
                "rows": rows,
                "format": "flattened_dict"
            }
            
        except Exception as e:
            logger.error(f"Failed to convert dict to table: {str(e)}")
            raise
    
    def _flatten_dict(self, data: Dict[str, Any], prefix: str = "") -> Dict[str, Any]:
        """扁平化嵌套字典"""
        result = {}
        
        for key, value in data.items():
            new_key = f"{prefix}.{key}" if prefix else key
            
            if isinstance(value, dict):
                result.update(self._flatten_dict(value, new_key))
            elif isinstance(value, list):
                for i, item in enumerate(value):
                    if isinstance(item, dict):
                        result.update(self._flatten_dict(item, f"{new_key}[{i}]"))
                    else:
                        result[f"{new_key}[{i}]"] = item
            else:
                result[new_key] = value
        
        return result
    
    def _detect_header_rows(self, data: List[List[Any]]) -> int:
        """检测表头行数"""
        if not data:
            return 0
        
        header_rows = 1
        
        # 检查前几行是否包含表头特征
        for i in range(1, min(5, len(data))):
            row = data[i]
            
            # 检查是否包含非数字的字符串（可能是表头）
            has_text_header = any(
                isinstance(cell, str) and 
                cell.strip() and 
                not str(cell).replace('.', '').replace('-', '').isdigit()
                for cell in row
            )
            
            if has_text_header:
                header_rows += 1
            else:
                break
        
        return header_rows
    
    def _merge_header_rows(self, header_rows: List[List[Any]]) -> List[str]:
        """合并多行表头"""
        if not header_rows:
            return []
        
        if len(header_rows) == 1:
            return [str(cell) if cell is not None else "" for cell in header_rows[0]]
        
        # 合并多行表头
        merged_headers = []
        max_cols = max(len(row) for row in header_rows)
        
        for col in range(max_cols):
            header_parts = []
            
            for row in header_rows:
                if col < len(row) and row[col] is not None:
                    header_parts.append(str(row[col]).strip())
            
            merged_header = " - ".join(filter(None, header_parts))
            merged_headers.append(merged_header)
        
        return merged_headers
    
    async def _check_data_quality(self, parsed_data: Dict[str, Any]) -> Dict[str, Any]:
        """检查数据质量"""
        try:
            return await self.quality_checker.check_quality(parsed_data)
        except Exception as e:
            logger.error(f"Failed to check data quality: {str(e)}")
            return {
                "overall_score": 0.0,
                "quality_level": DataQualityLevel.POOR,
                "issues": [str(e)],
                "warnings": []
            }
    
    async def _apply_field_mappings(
        self, 
        parsed_data: Dict[str, Any], 
        field_mappings: List[FieldMapping]
    ) -> Dict[str, Any]:
        """应用字段映射"""
        try:
            return await self.field_mapper.apply_mappings(parsed_data, field_mappings)
        except Exception as e:
            logger.error(f"Failed to apply field mappings: {str(e)}")
            raise
    
    async def _validate_data(
        self, 
        mapped_data: Dict[str, Any], 
        field_mappings: List[FieldMapping]
    ) -> Dict[str, Any]:
        """验证数据"""
        try:
            validation_result = {
                "valid": True,
                "errors": [],
                "warnings": []
            }
            
            # 应用验证规则
            for mapping in field_mappings:
                if mapping.validation_rule:
                    field_validation = await self._validate_field(
                        mapped_data, 
                        mapping.target_field, 
                        mapping.validation_rule
                    )
                    
                    if not field_validation["valid"]:
                        validation_result["valid"] = False
                        validation_result["errors"].extend(field_validation["errors"])
                    
                    validation_result["warnings"].extend(field_validation["warnings"])
            
            return validation_result
            
        except Exception as e:
            logger.error(f"Failed to validate data: {str(e)}")
            return {
                "valid": False,
                "errors": [str(e)],
                "warnings": []
            }
    
    async def _validate_field(
        self, 
        data: Dict[str, Any], 
        field_name: str, 
        validation_rule: str
    ) -> Dict[str, Any]:
        """验证单个字段"""
        try:
            # 这里实现具体的字段验证逻辑
            # 支持范围验证、格式验证、枚举验证等
            
            result = {
                "valid": True,
                "errors": [],
                "warnings": []
            }
            
            # 解析验证规则
            rule_parts = validation_rule.split(":")
            rule_type = rule_parts[0]
            rule_value = rule_parts[1] if len(rule_parts) > 1 else None
            
            # 获取字段数据
            field_data = self._extract_field_data(data, field_name)
            
            if rule_type == "range":
                # 范围验证
                min_val, max_val = map(float, rule_value.split(","))
                for value in field_data:
                    if isinstance(value, (int, float)) and not (min_val <= value <= max_val):
                        result["valid"] = False
                        result["errors"].append(f"字段 {field_name} 的值 {value} 超出范围 [{min_val}, {max_val}]")
            
            elif rule_type == "format":
                # 格式验证
                pattern = re.compile(rule_value)
                for value in field_data:
                    if isinstance(value, str) and not pattern.match(value):
                        result["valid"] = False
                        result["errors"].append(f"字段 {field_name} 的值 {value} 不符合格式 {rule_value}")
            
            elif rule_type == "enum":
                # 枚举验证
                allowed_values = rule_value.split(",")
                for value in field_data:
                    if str(value) not in allowed_values:
                        result["valid"] = False
                        result["errors"].append(f"字段 {field_name} 的值 {value} 不在允许的枚举值中: {allowed_values}")
            
            return result
            
        except Exception as e:
            logger.error(f"Failed to validate field: {str(e)}")
            return {
                "valid": False,
                "errors": [str(e)],
                "warnings": []
            }
    
    def _extract_field_data(self, data: Dict[str, Any], field_name: str) -> List[Any]:
        """提取字段数据"""
        try:
            if "sheets" in data:
                # Excel数据
                field_data = []
                for sheet_name, sheet_data in data["sheets"].items():
                    if "rows" in sheet_data:
                        headers = sheet_data.get("headers", [])
                        if field_name in headers:
                            field_index = headers.index(field_name)
                            for row in sheet_data["rows"]:
                                if field_index < len(row):
                                    field_data.append(row[field_index])
                return field_data
            
            elif "rows" in data:
                # CSV或JSON数据
                field_data = []
                headers = data.get("headers", [])
                if field_name in headers:
                    field_index = headers.index(field_name)
                    for row in data["rows"]:
                        if field_index < len(row):
                            field_data.append(row[field_index])
                return field_data
            
            return []
            
        except Exception as e:
            logger.error(f"Failed to extract field data: {str(e)}")
            return []
    
    async def _clean_data(
        self, 
        mapped_data: Dict[str, Any], 
        validation_result: Dict[str, Any]
    ) -> Dict[str, Any]:
        """清洗数据"""
        try:
            return await self.data_transformer.clean_data(mapped_data, validation_result)
        except Exception as e:
            logger.error(f"Failed to clean data: {str(e)}")
            raise
    
    async def _load_data_to_target(
        self, 
        cleaned_data: Dict[str, Any], 
        target_table: str, 
        import_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """加载数据到目标表"""
        try:
            # 这里实现数据加载到数据库的逻辑
            # 支持批量插入、更新、合并等操作
            
            load_mode = import_config.get("load_mode", "insert")  # insert, update, merge
            
            if load_mode == "insert":
                return await self._insert_data(cleaned_data, target_table)
            elif load_mode == "update":
                return await self._update_data(cleaned_data, target_table, import_config)
            elif load_mode == "merge":
                return await self._merge_data(cleaned_data, target_table, import_config)
            else:
                raise ValueError(f"不支持的加载模式: {load_mode}")
                
        except Exception as e:
            logger.error(f"Failed to load data to target: {str(e)}")
            raise
    
    async def _insert_data(self, data: Dict[str, Any], target_table: str) -> Dict[str, Any]:
        """插入数据"""
        try:
            # 实现批量插入逻辑
            successful_count = 0
            failed_count = 0
            
            # 这里调用数据库服务进行批量插入
            # result = await self.db_service.bulk_insert(target_table, data)
            
            return {
                "success": True,
                "successful_count": successful_count,
                "failed_count": failed_count,
                "load_mode": "insert"
            }
            
        except Exception as e:
            logger.error(f"Failed to insert data: {str(e)}")
            raise
    
    async def _update_data(
        self, 
        data: Dict[str, Any], 
        target_table: str, 
        import_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """更新数据"""
        try:
            # 实现更新逻辑
            update_key = import_config.get("update_key", "id")
            
            successful_count = 0
            failed_count = 0
            
            # 这里调用数据库服务进行批量更新
            # result = await self.db_service.bulk_update(target_table, data, update_key)
            
            return {
                "success": True,
                "successful_count": successful_count,
                "failed_count": failed_count,
                "load_mode": "update"
            }
            
        except Exception as e:
            logger.error(f"Failed to update data: {str(e)}")
            raise
    
    async def _merge_data(
        self, 
        data: Dict[str, Any], 
        target_table: str, 
        import_config: Dict[str, Any]
    ) -> Dict[str, Any]:
        """合并数据"""
        try:
            # 实现合并逻辑（UPSERT）
            merge_key = import_config.get("merge_key", "id")
            
            successful_count = 0
            failed_count = 0
            
            # 这里调用数据库服务进行批量合并
            # result = await self.db_service.bulk_merge(target_table, data, merge_key)
            
            return {
                "success": True,
                "successful_count": successful_count,
                "failed_count": failed_count,
                "load_mode": "merge"
            }
            
        except Exception as e:
            logger.error(f"Failed to merge data: {str(e)}")
            raise
    
    def _load_validation_rules(self) -> List[DataValidationRule]:
        """加载验证规则"""
        return [
            DataValidationRule("email", "format", r"^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$", "邮箱格式不正确"),
            DataValidationRule("phone", "format", r"^1[3-9]\d{9}$", "手机号格式不正确"),
            DataValidationRule("amount", "range", "0,999999999", "金额超出有效范围"),
            DataValidationRule("status", "enum", "active,inactive,pending", "状态值不在允许范围内")
        ]

# ==================== 辅助类 ====================

class DataQualityChecker:
    """数据质量检查器"""
    
    async def check_quality(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """检查数据质量"""
        try:
            issues = []
            warnings = []
            
            # 检查缺失值
            missing_issues = await self._check_missing_values(data)
            issues.extend(missing_issues)
            
            # 检查重复值
            duplicate_issues = await self._check_duplicates(data)
            issues.extend(duplicate_issues)
            
            # 检查数据一致性
            consistency_issues = await self._check_consistency(data)
            issues.extend(consistency_issues)
            
            # 检查数据格式
            format_issues = await self._check_formats(data)
            warnings.extend(format_issues)
            
            # 计算质量分数
            total_checks = 4
            failed_checks = len(issues)
            quality_score = max(0, (total_checks - failed_checks) / total_checks)
            
            # 确定质量等级
            if quality_score >= 0.95:
                quality_level = DataQualityLevel.EXCELLENT
            elif quality_score >= 0.85:
                quality_level = DataQualityLevel.GOOD
            elif quality_score >= 0.70:
                quality_level = DataQualityLevel.FAIR
            else:
                quality_level = DataQualityLevel.POOR
            
            return {
                "overall_score": quality_score,
                "quality_level": quality_level,
                "issues": issues,
                "warnings": warnings
            }
            
        except Exception as e:
            logger.error(f"Failed to check data quality: {str(e)}")
            return {
                "overall_score": 0.0,
                "quality_level": DataQualityLevel.POOR,
                "issues": [str(e)],
                "warnings": []
            }
    
    async def _check_missing_values(self, data: Dict[str, Any]) -> List[str]:
        """检查缺失值"""
        issues = []
        
        try:
            if "sheets" in data:
                for sheet_name, sheet_data in data["sheets"].items():
                    rows = sheet_data.get("rows", [])
                    headers = sheet_data.get("headers", [])
                    
                    for i, row in enumerate(rows):
                        for j, cell in enumerate(row):
                            if cell is None or (isinstance(cell, str) and not cell.strip()):
                                if j < len(headers):
                                    issues.append(f"工作表 {sheet_name} 第 {i+1} 行字段 {headers[j]} 存在缺失值")
            
            elif "rows" in data:
                rows = data["rows"]
                headers = data.get("headers", [])
                
                for i, row in enumerate(rows):
                    for j, cell in enumerate(row):
                        if cell is None or (isinstance(cell, str) and not cell.strip()):
                            if j < len(headers):
                                issues.append(f"第 {i+1} 行字段 {headers[j]} 存在缺失值")
            
            return issues
            
        except Exception as e:
            logger.error(f"Failed to check missing values: {str(e)}")
            return [str(e)]
    
    async def _check_duplicates(self, data: Dict[str, Any]) -> List[str]:
        """检查重复值"""
        issues = []
        
        try:
            if "sheets" in data:
                for sheet_name, sheet_data in data["sheets"].items():
                    rows = sheet_data.get("rows", [])
                    
                    # 检查完全重复的行
                    seen_rows = set()
                    for i, row in enumerate(rows):
                        row_tuple = tuple(row)
                        if row_tuple in seen_rows:
                            issues.append(f"工作表 {sheet_name} 第 {i+1} 行与之前的行完全重复")
                        else:
                            seen_rows.add(row_tuple)
            
            elif "rows" in data:
                rows = data["rows"]
                
                # 检查完全重复的行
                seen_rows = set()
                for i, row in enumerate(rows):
                    row_tuple = tuple(row)
                    if row_tuple in seen_rows:
                        issues.append(f"第 {i+1} 行与之前的行完全重复")
                    else:
                        seen_rows.add(row_tuple)
            
            return issues
            
        except Exception as e:
            logger.error(f"Failed to check duplicates: {str(e)}")
            return [str(e)]
    
    async def _check_consistency(self, data: Dict[str, Any]) -> List[str]:
        """检查数据一致性"""
        issues = []
        
        try:
            # 这里可以实现更复杂的一致性检查
            # 例如：日期格式一致性、数值范围一致性等
            
            return issues
            
        except Exception as e:
            logger.error(f"Failed to check consistency: {str(e)}")
            return [str(e)]
    
    async def _check_formats(self, data: Dict[str, Any]) -> List[str]:
        """检查数据格式"""
        warnings = []
        
        try:
            # 这里可以实现格式检查
            # 例如：日期格式、数字格式等
            
            return warnings
            
        except Exception as e:
            logger.error(f"Failed to check formats: {str(e)}")
            return [str(e)]

class FieldMapper:
    """字段映射器"""
    
    async def apply_mappings(
        self, 
        parsed_data: Dict[str, Any], 
        field_mappings: List[FieldMapping]
    ) -> Dict[str, Any]:
        """应用字段映射"""
        try:
            mapped_data = parsed_data.copy()
            
            # 创建映射字典
            mapping_dict = {mapping.source_field: mapping for mapping in field_mappings}
            
            # 应用映射
            if "sheets" in mapped_data:
                for sheet_name, sheet_data in mapped_data["sheets"].items():
                    headers = sheet_data.get("headers", [])
                    rows = sheet_data.get("rows", [])
                    
                    # 映射表头
                    mapped_headers = []
                    for header in headers:
                        if header in mapping_dict:
                            mapped_headers.append(mapping_dict[header].target_field)
                        else:
                            mapped_headers.append(header)
                    
                    # 应用数据转换
                    mapped_rows = []
                    for row in rows:
                        mapped_row = []
                        for i, cell in enumerate(row):
                            if i < len(headers) and headers[i] in mapping_dict:
                                mapping = mapping_dict[headers[i]]
                                transformed_cell = await self._transform_cell(cell, mapping)
                                mapped_row.append(transformed_cell)
                            else:
                                mapped_row.append(cell)
                        mapped_rows.append(mapped_row)
                    
                    sheet_data["headers"] = mapped_headers
                    sheet_data["rows"] = mapped_rows
            
            elif "rows" in mapped_data:
                headers = mapped_data.get("headers", [])
                rows = mapped_data.get("rows", [])
                
                # 映射表头
                mapped_headers = []
                for header in headers:
                    if header in mapping_dict:
                        mapped_headers.append(mapping_dict[header].target_field)
                    else:
                        mapped_headers.append(header)
                
                # 应用数据转换
                mapped_rows = []
                for row in rows:
                    mapped_row = []
                    for i, cell in enumerate(row):
                        if i < len(headers) and headers[i] in mapping_dict:
                            mapping = mapping_dict[headers[i]]
                            transformed_cell = await self._transform_cell(cell, mapping)
                            mapped_row.append(transformed_cell)
                        else:
                            mapped_row.append(cell)
                    mapped_rows.append(mapped_row)
                
                mapped_data["headers"] = mapped_headers
                mapped_data["rows"] = mapped_rows
            
            return mapped_data
            
        except Exception as e:
            logger.error(f"Failed to apply field mappings: {str(e)}")
            raise
    
    async def _transform_cell(self, cell: Any, mapping: FieldMapping) -> Any:
        """转换单元格数据"""
        try:
            if mapping.transformation_rule:
                # 应用转换规则
                return await self._apply_transformation_rule(cell, mapping.transformation_rule)
            else:
                # 基本类型转换
                return await self._convert_data_type(cell, mapping.data_type)
                
        except Exception as e:
            logger.error(f"Failed to transform cell: {str(e)}")
            return cell
    
    async def _apply_transformation_rule(self, cell: Any, rule: str) -> Any:
        """应用转换规则"""
        try:
            # 这里可以实现各种转换规则
            # 例如：字符串替换、数学运算、日期格式转换等
            
            if rule.startswith("replace:"):
                # 字符串替换
                old, new = rule.split(":", 2)[1:]
                return str(cell).replace(old, new)
            
            elif rule.startswith("format_date:"):
                # 日期格式转换
                target_format = rule.split(":", 1)[1]
                from datetime import datetime
                if isinstance(cell, str):
                    # 尝试解析日期
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"]:
                        try:
                            dt = datetime.strptime(cell, fmt)
                            return dt.strftime(target_format)
                        except ValueError:
                            continue
                return cell
            
            else:
                return cell
                
        except Exception as e:
            logger.error(f"Failed to apply transformation rule: {str(e)}")
            return cell
    
    async def _convert_data_type(self, cell: Any, data_type: str) -> Any:
        """转换数据类型"""
        try:
            if data_type == "string":
                return str(cell) if cell is not None else ""
            elif data_type == "integer":
                return int(float(cell)) if cell is not None and str(cell).strip() else None
            elif data_type == "float":
                return float(cell) if cell is not None and str(cell).strip() else None
            elif data_type == "boolean":
                if isinstance(cell, bool):
                    return cell
                elif isinstance(cell, str):
                    return cell.lower() in ["true", "1", "yes", "是"]
                elif isinstance(cell, (int, float)):
                    return bool(cell)
                else:
                    return False
            elif data_type == "date":
                from datetime import datetime
                if isinstance(cell, str) and cell.strip():
                    # 尝试解析日期
                    for fmt in ["%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y", "%Y%m%d"]:
                        try:
                            return datetime.strptime(cell, fmt).date()
                        except ValueError:
                            continue
                return cell
            else:
                return cell
                
        except Exception as e:
            logger.error(f"Failed to convert data type: {str(e)}")
            return cell

class DataTransformer:
    """数据转换器"""
    
    async def clean_data(
        self, 
        mapped_data: Dict[str, Any], 
        validation_result: Dict[str, Any]
    ) -> Dict[str, Any]:
        """清洗数据"""
        try:
            cleaned_data = mapped_data.copy()
            
            # 处理缺失值
            cleaned_data = await self._handle_missing_values(cleaned_data)
            
            # 处理异常值
            cleaned_data = await self._handle_outliers(cleaned_data)
            
            # 标准化数据
            cleaned_data = await self._standardize_data(cleaned_data)
            
            return cleaned_data
            
        except Exception as e:
            logger.error(f"Failed to clean data: {str(e)}")
            raise
    
    async def _handle_missing_values(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """处理缺失值"""
        try:
            if "sheets" in data:
                for sheet_name, sheet_data in data["sheets"].items():
                    rows = sheet_data.get("rows", [])
                    
                    # 删除完全空的行
                    cleaned_rows = []
                    for row in rows:
                        if any(cell is not None and str(cell).strip() for cell in row):
                            cleaned_rows.append(row)
                    
                    sheet_data["rows"] = cleaned_rows
            
            elif "rows" in data:
                rows = data["rows"]
                
                # 删除完全空的行
                cleaned_rows = []
                for row in rows:
                    if any(cell is not None and str(cell).strip() for cell in row):
                        cleaned_rows.append(row)
                
                data["rows"] = cleaned_rows
            
            return data
            
        except Exception as e:
            logger.error(f"Failed to handle missing values: {str(e)}")
            return data
    
    async def _handle_outliers(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """处理异常值"""
        try:
            # 这里可以实现异常值检测和处理
            # 例如：使用IQR方法、Z-score方法等
            
            return data
            
        except Exception as e:
            logger.error(f"Failed to handle outliers: {str(e)}")
            return data
    
    async def _standardize_data(self, data: Dict[str, Any]) -> Dict[str, Any]:
        """标准化数据"""
        try:
            # 这里可以实现数据标准化
            # 例如：字符串去空格、数字格式化等
            
            if "sheets" in data:
                for sheet_name, sheet_data in data["sheets"].items():
                    rows = sheet_data.get("rows", [])
                    
                    standardized_rows = []
                    for row in rows:
                        standardized_row = []
                        for cell in row:
                            if isinstance(cell, str):
                                standardized_row.append(cell.strip())
                            else:
                                standardized_row.append(cell)
                        standardized_rows.append(standardized_row)
                    
                    sheet_data["rows"] = standardized_rows
            
            elif "rows" in data:
                rows = data["rows"]
                
                standardized_rows = []
                for row in rows:
                    standardized_row = []
                    for cell in row:
                        if isinstance(cell, str):
                            standardized_row.append(cell.strip())
                        else:
                            standardized_row.append(cell)
                    standardized_rows.append(standardized_row)
                
                data["rows"] = standardized_rows
            
            return data
            
        except Exception as e:
            logger.error(f"Failed to standardize data: {str(e)}")
            return data

