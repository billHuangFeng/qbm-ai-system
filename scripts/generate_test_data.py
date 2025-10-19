#!/usr/bin/env python3
"""
测试数据生成脚本 - QBM AI System
"""
import random
import csv
import json
from datetime import datetime, timedelta
from pathlib import Path

class TestDataGenerator:
    def __init__(self):
        self.output_dir = Path("test_data")
        self.output_dir.mkdir(exist_ok=True)
        
        # 测试数据模板
        self.industries = ["科技", "金融", "教育", "医疗", "制造", "零售", "房地产", "咨询"]
        self.regions = ["北京", "上海", "广州", "深圳", "杭州", "南京", "成都", "武汉"]
        self.product_categories = ["软件", "硬件", "服务", "咨询", "培训", "维护"]
        self.contract_types = ["销售合同", "服务合同", "维护合同", "培训合同"]
        
    def generate_customers(self, count=100):
        """生成客户测试数据"""
        self.log(f"生成 {count} 个客户测试数据...")
        
        customers = []
        for i in range(count):
            customer = {
                "name": f"测试客户{i+1:03d}",
                "contact_person": f"联系人{i+1}",
                "contact_email": f"customer{i+1:03d}@example.com",
                "contact_phone": f"138{random.randint(10000000, 99999999)}",
                "industry": random.choice(self.industries),
                "region": random.choice(self.regions),
                "address": f"{random.choice(self.regions)}市{random.randint(1, 10)}区测试路{i+1}号",
                "description": f"这是第{i+1}个测试客户，用于系统测试。"
            }
            customers.append(customer)
        
        # 保存为CSV
        csv_file = self.output_dir / "test_customers.csv"
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=customers[0].keys())
            writer.writeheader()
            writer.writerows(customers)
        
        # 保存为JSON
        json_file = self.output_dir / "test_customers.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(customers, f, ensure_ascii=False, indent=2)
        
        self.log(f"✅ 客户数据已生成: {csv_file}, {json_file}")
        return customers
    
    def generate_products(self, count=50):
        """生成产品测试数据"""
        self.log(f"生成 {count} 个产品测试数据...")
        
        products = []
        for i in range(count):
            price = round(random.uniform(100, 10000), 2)
            product = {
                "name": f"测试产品{i+1:03d}",
                "description": f"这是第{i+1}个测试产品，用于系统测试。产品具有高质量和良好的性能。",
                "price": price,
                "category": random.choice(self.product_categories),
                "stock_quantity": random.randint(10, 1000)
            }
            products.append(product)
        
        # 保存为CSV
        csv_file = self.output_dir / "test_products.csv"
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=products[0].keys())
            writer.writeheader()
            writer.writerows(products)
        
        # 保存为JSON
        json_file = self.output_dir / "test_products.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(products, f, ensure_ascii=False, indent=2)
        
        self.log(f"✅ 产品数据已生成: {csv_file}, {json_file}")
        return products
    
    def generate_orders(self, customers, products, count=200):
        """生成订单测试数据"""
        self.log(f"生成 {count} 个订单测试数据...")
        
        orders = []
        for i in range(count):
            customer = random.choice(customers)
            product = random.choice(products)
            quantity = random.randint(1, 10)
            total_amount = product["price"] * quantity
            
            # 生成随机日期（过去30天内）
            order_date = datetime.now() - timedelta(days=random.randint(1, 30))
            
            order = {
                "customer_name": customer["name"],
                "product_name": product["name"],
                "quantity": quantity,
                "unit_price": product["price"],
                "total_amount": total_amount,
                "order_date": order_date.strftime("%Y-%m-%d"),
                "status": random.choice(["已完成", "进行中", "已取消"])
            }
            orders.append(order)
        
        # 保存为CSV
        csv_file = self.output_dir / "test_orders.csv"
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=orders[0].keys())
            writer.writeheader()
            writer.writerows(orders)
        
        # 保存为JSON
        json_file = self.output_dir / "test_orders.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(orders, f, ensure_ascii=False, indent=2)
        
        self.log(f"✅ 订单数据已生成: {csv_file}, {json_file}")
        return orders
    
    def generate_contracts(self, customers, count=80):
        """生成合同测试数据"""
        self.log(f"生成 {count} 个合同测试数据...")
        
        contracts = []
        for i in range(count):
            customer = random.choice(customers)
            start_date = datetime.now() - timedelta(days=random.randint(1, 365))
            end_date = start_date + timedelta(days=random.randint(30, 365))
            amount = round(random.uniform(10000, 500000), 2)
            
            contract = {
                "contract_number": f"CT{i+1:06d}",
                "customer_name": customer["name"],
                "contract_type": random.choice(self.contract_types),
                "start_date": start_date.strftime("%Y-%m-%d"),
                "end_date": end_date.strftime("%Y-%m-%d"),
                "amount": amount,
                "status": random.choice(["有效", "已到期", "已终止"])
            }
            contracts.append(contract)
        
        # 保存为CSV
        csv_file = self.output_dir / "test_contracts.csv"
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=contracts[0].keys())
            writer.writeheader()
            writer.writerows(contracts)
        
        # 保存为JSON
        json_file = self.output_dir / "test_contracts.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(contracts, f, ensure_ascii=False, indent=2)
        
        self.log(f"✅ 合同数据已生成: {csv_file}, {json_file}")
        return contracts
    
    def generate_financial_records(self, count=150):
        """生成财务记录测试数据"""
        self.log(f"生成 {count} 个财务记录测试数据...")
        
        financial_records = []
        for i in range(count):
            record_date = datetime.now() - timedelta(days=random.randint(1, 90))
            amount = round(random.uniform(1000, 100000), 2)
            
            record = {
                "type": random.choice(["收入", "支出"]),
                "amount": amount,
                "description": f"测试财务记录{i+1}",
                "date": record_date.strftime("%Y-%m-%d"),
                "category": random.choice(["销售", "采购", "运营", "投资", "其他"])
            }
            financial_records.append(record)
        
        # 保存为CSV
        csv_file = self.output_dir / "test_financials.csv"
        with open(csv_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=financial_records[0].keys())
            writer.writeheader()
            writer.writerows(financial_records)
        
        # 保存为JSON
        json_file = self.output_dir / "test_financials.json"
        with open(json_file, 'w', encoding='utf-8') as f:
            json.dump(financial_records, f, ensure_ascii=False, indent=2)
        
        self.log(f"✅ 财务记录已生成: {csv_file}, {json_file}")
        return financial_records
    
    def generate_excel_file(self, customers, products, orders, contracts, financials):
        """生成Excel测试文件"""
        self.log("生成Excel测试文件...")
        
        try:
            import openpyxl
            from openpyxl import Workbook
            
            wb = Workbook()
            
            # 删除默认工作表
            wb.remove(wb.active)
            
            # 创建各个工作表
            sheets_data = [
                ("客户", customers),
                ("产品", products),
                ("订单", orders),
                ("合同", contracts),
                ("财务", financials)
            ]
            
            for sheet_name, data in sheets_data:
                ws = wb.create_sheet(title=sheet_name)
                
                if data:
                    # 写入表头
                    headers = list(data[0].keys())
                    for col, header in enumerate(headers, 1):
                        ws.cell(row=1, column=col, value=header)
                    
                    # 写入数据
                    for row, record in enumerate(data, 2):
                        for col, value in enumerate(record.values(), 1):
                            ws.cell(row=row, column=col, value=value)
            
            # 保存文件
            excel_file = self.output_dir / "test_data.xlsx"
            wb.save(excel_file)
            self.log(f"✅ Excel文件已生成: {excel_file}")
            
        except ImportError:
            self.log("⚠️  openpyxl未安装，跳过Excel文件生成", "WARNING")
    
    def log(self, message, level="INFO"):
        """记录日志"""
        timestamp = datetime.now().strftime("%H:%M:%S")
        print(f"[{timestamp}] [{level}] {message}")
    
    def generate_all(self):
        """生成所有测试数据"""
        self.log("🚀 开始生成测试数据...")
        self.log("=" * 50)
        
        # 生成基础数据
        customers = self.generate_customers(100)
        products = self.generate_products(50)
        
        # 生成关联数据
        orders = self.generate_orders(customers, products, 200)
        contracts = self.generate_contracts(customers, 80)
        financials = self.generate_financial_records(150)
        
        # 生成Excel文件
        self.generate_excel_file(customers, products, orders, contracts, financials)
        
        self.log("\n" + "=" * 50)
        self.log("📊 测试数据生成完成:")
        self.log(f"客户数据: {len(customers)} 条")
        self.log(f"产品数据: {len(products)} 条")
        self.log(f"订单数据: {len(orders)} 条")
        self.log(f"合同数据: {len(contracts)} 条")
        self.log(f"财务数据: {len(financials)} 条")
        self.log(f"输出目录: {self.output_dir.absolute()}")
        
        self.log("\n📁 生成的文件:")
        for file_path in self.output_dir.glob("*"):
            self.log(f"  - {file_path.name}")
        
        self.log("\n💡 使用说明:")
        self.log("1. 可以通过数据导入功能导入这些测试数据")
        self.log("2. 支持CSV、JSON、Excel格式")
        self.log("3. 数据已包含合理的关联关系")
        self.log("4. 可用于功能测试和演示")

def main():
    """主函数"""
    print("QBM AI System 测试数据生成工具")
    print("=" * 50)
    
    generator = TestDataGenerator()
    generator.generate_all()

if __name__ == "__main__":
    main()


